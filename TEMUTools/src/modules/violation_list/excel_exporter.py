from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from typing import List, Dict, Any
import logging

logger = logging.getLogger(__name__)

class ViolationListExcelExporter:
    """违规商品列表Excel导出器"""
    
    def __init__(self):
        self.columns = [
            'SPUID', '商品ID', '商品名称', '违规类型', '违规描述', '处罚类型',
            '处罚描述', '违规详情', '整改建议', '处罚开始时间', '处罚结束时间',
            '申诉状态', '申诉次数', '最大申诉次数'
        ]
    
    def format_timestamp(self, timestamp: int) -> str:
        """格式化时间戳为可读时间"""
        if not timestamp:
            return ''
        try:
            return datetime.fromtimestamp(timestamp).strftime('%Y-%m-%d %H:%M:%S')
        except:
            return ''
    
    def format_appeal_status(self, status: int) -> str:
        """格式化申诉状态"""
        status_map = {
            0: '未申诉',
            1: '申诉中',
            2: '申诉通过',
            3: '申诉驳回'
        }
        return status_map.get(status, '未知状态')
    
    def export_to_excel(self, data: List[Dict[str, Any]], file_path: str) -> bool:
        """导出数据到Excel文件
        
        Args:
            data: 违规商品数据列表
            file_path: 导出文件路径
            
        Returns:
            bool: 是否导出成功
        """
        try:
            # 创建工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "违规商品列表"
            
            # 设置样式
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 写入表头
            for col, header in enumerate(self.columns, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
                # 设置列宽
                ws.column_dimensions[get_column_letter(col)].width = 20
            
            # 准备数据
            row_num = 2
            for product in data:
                # 获取处罚详情
                punish_details = product.get('punish_detail_list', [])
                if not punish_details:
                    # 如果没有处罚详情，添加基本信息
                    values = [
                        product.get('spu_id', ''),
                        product.get('goods_id', ''),
                        product.get('goods_name', ''),
                        product.get('leaf_reason_name', ''),
                        product.get('violation_desc', ''),
                        '', '', '', '', '', '',
                        self.format_appeal_status(product.get('appeal_status', 0)),
                        '', ''
                    ]
                    
                    # 写入数据
                    for col, value in enumerate(values, 1):
                        cell = ws.cell(row=row_num, column=col, value=value)
                        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                        cell.border = border
                    row_num += 1
                    continue
                
                # 处理每个处罚详情
                for detail in punish_details:
                    # 获取违规详情
                    illegal_details = detail.get('illegal_detail', [])
                    illegal_desc = '; '.join([
                        f"{illegal.get('title', '')}: {illegal.get('value', '')}"
                        for illegal in illegal_details
                    ])
                    
                    values = [
                        product.get('spu_id', ''),
                        product.get('goods_id', ''),
                        product.get('goods_name', ''),
                        product.get('leaf_reason_name', ''),
                        product.get('violation_desc', ''),
                        detail.get('punish_appeal_type', ''),
                        detail.get('punish_infect_desc', ''),
                        illegal_desc,
                        detail.get('rectification_suggestion', ''),
                        self.format_timestamp(detail.get('start_time')),
                        self.format_timestamp(detail.get('plan_end_time')),
                        self.format_appeal_status(detail.get('appeal_status', 0)),
                        detail.get('now_appeal_time', 0),
                        detail.get('max_appeal_time', 0)
                    ]
                    
                    # 写入数据
                    for col, value in enumerate(values, 1):
                        cell = ws.cell(row=row_num, column=col, value=value)
                        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                        cell.border = border
                    row_num += 1
            
            # 保存文件
            wb.save(file_path)
            
            logger.info(f"成功导出Excel文件: {file_path}")
            return True
            
        except Exception as e:
            logger.error(f"导出Excel文件时出错: {str(e)}")
            return False 