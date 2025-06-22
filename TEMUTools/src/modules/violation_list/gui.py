import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import json
import os
import threading
from datetime import datetime
from .crawler import ViolationListCrawler
from .excel_exporter import ViolationListExcelExporter
import logging
from ..system_config.config import SystemConfig

logger = logging.getLogger(__name__)

class ViolationListTab(ttk.Frame):
    def __init__(self, parent):
        super().__init__(parent)
        self.crawler = ViolationListCrawler()
        self.excel_exporter = ViolationListExcelExporter()
        self.config = SystemConfig()
        
        # 设置配置文件路径
        self.config_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'config')
        self.config_file = os.path.join(self.config_dir, 'violation_config.json')
        
        # 确保配置目录存在
        os.makedirs(self.config_dir, exist_ok=True)
        
        # 初始化默认值
        self.last_pages = '1'
        self.last_page_size = '100'
        
        self.setup_ui()
        self.load_config()
        self.setup_logging()
        
    def setup_ui(self):
        """设置用户界面"""
        # 创建主框架
        main_frame = ttk.Frame(self, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # 创建输入字段
        self.create_input_fields(main_frame)
        
        # 创建进度条
        self.create_progress_bar(main_frame)
        
        # 创建日志显示区域
        self.create_log_area(main_frame)
        
        # 创建按钮
        self.create_buttons(main_frame)
        
        # 配置网格权重
        self.columnconfigure(0, weight=1)
        self.rowconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)
        
    def create_input_fields(self, parent):
        """创建输入字段"""
        # 创建一个子框架来容纳输入字段
        input_frame = ttk.Frame(parent)
        input_frame.grid(row=0, column=0, columnspan=2, pady=5, sticky=tk.W)
        
        # 起始页输入
        ttk.Label(input_frame, text="起始页:").grid(row=0, column=0, sticky=tk.W, padx=(0, 5))
        self.start_page_var = tk.StringVar(value="1")
        self.start_page_entry = ttk.Entry(input_frame, textvariable=self.start_page_var, width=10)
        self.start_page_entry.grid(row=0, column=1, sticky=tk.W, padx=(0, 20))
        
        # 结束页输入
        ttk.Label(input_frame, text="结束页:").grid(row=0, column=2, sticky=tk.W, padx=(0, 5))
        self.end_page_var = tk.StringVar(value="1")
        self.end_page_entry = ttk.Entry(input_frame, textvariable=self.end_page_var, width=10)
        self.end_page_entry.grid(row=0, column=3, sticky=tk.W, padx=(0, 20))
        
        # 每页数量输入
        ttk.Label(input_frame, text="每页数量:").grid(row=0, column=4, sticky=tk.W, padx=(0, 5))
        self.page_size_var = tk.StringVar(value="100")
        self.page_size_entry = ttk.Entry(input_frame, textvariable=self.page_size_var, width=10)
        self.page_size_entry.grid(row=0, column=5, sticky=tk.W)
        
        # 配置输入框架的网格权重，使其左对齐
        input_frame.columnconfigure(0, weight=0)
        input_frame.columnconfigure(1, weight=0)
        input_frame.columnconfigure(2, weight=0)
        input_frame.columnconfigure(3, weight=0)
        input_frame.columnconfigure(4, weight=0)
        input_frame.columnconfigure(5, weight=0)
        
        # 配置父框架的网格权重
        parent.columnconfigure(0, weight=1)
        parent.columnconfigure(1, weight=1)
        
    def create_progress_bar(self, parent):
        """创建进度条"""
        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(
            parent,
            variable=self.progress_var,
            maximum=100
        )
        self.progress_bar.grid(row=1, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=5)
        
    def create_log_area(self, parent):
        """创建日志显示区域"""
        log_frame = ttk.LabelFrame(parent, text="操作日志", padding="5")
        log_frame.grid(row=2, column=0, columnspan=2, sticky=(tk.W, tk.E, tk.N, tk.S), pady=5)
        
        # 创建日志文本框
        self.log_text = tk.Text(log_frame, height=15, width=80, state='disabled')
        self.log_text.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # 添加滚动条
        scrollbar = ttk.Scrollbar(log_frame, orient=tk.VERTICAL, command=self.log_text.yview)
        scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S))
        self.log_text['yscrollcommand'] = scrollbar.set
        
        log_frame.columnconfigure(0, weight=1)
        log_frame.rowconfigure(0, weight=1)
        
    def create_buttons(self, parent):
        """创建按钮"""
        # 创建按钮框架
        button_frame = ttk.Frame(parent)
        button_frame.grid(row=3, column=0, columnspan=2, pady=10)
        
        # 开始按钮
        self.start_button = ttk.Button(
            button_frame, 
            text="开始获取",
            command=self.start_crawling
        )
        self.start_button.pack(side=tk.LEFT, padx=5)
        
        # 停止按钮
        self.stop_button = ttk.Button(
            button_frame,
            text="停止",
            command=self.stop_crawling,
            state='disabled'
        )
        self.stop_button.pack(side=tk.LEFT, padx=5)
        
        # 导出Excel按钮
        self.export_button = ttk.Button(
            button_frame,
            text="导出Excel",
            command=self.export_to_excel,
            state='disabled'  # 初始状态为禁用
        )
        self.export_button.pack(side=tk.LEFT, padx=5)
        
        # 合并Excel按钮
        self.merge_button = ttk.Button(
            button_frame,
            text="合并Excel",
            command=self.merge_excel_files
        )
        self.merge_button.pack(side=tk.LEFT, padx=5)
        
        # 保存配置按钮
        self.save_config_button = ttk.Button(
            button_frame,
            text="保存配置",
            command=self.save_config
        )
        self.save_config_button.pack(side=tk.LEFT, padx=5)
        
        # 清空按钮
        self.clear_button = ttk.Button(
            button_frame,
            text="清空",
            command=self.clear_all
        )
        self.clear_button.pack(side=tk.LEFT, padx=5)
        
        # 存储当前数据
        self.violation_data = []
    
    def load_config(self):
        """加载配置"""
        try:
            if os.path.exists(self.config_file):
                with open(self.config_file, 'r', encoding='utf-8') as f:
                    config = json.load(f)
                    self.start_page_var.set(config.get('start_page', '1'))
                    self.end_page_var.set(config.get('end_page', '1'))
                    self.page_size_var.set(config.get('page_size', '100'))
        except Exception as e:
            self.logger.error(f"加载配置失败: {str(e)}")
            # 使用默认值
            self.start_page_var.set('1')
            self.end_page_var.set('1')
            self.page_size_var.set('100')
    
    def save_config(self):
        """保存配置"""
        try:
            # 确保配置目录存在
            os.makedirs(os.path.dirname(self.config_file), exist_ok=True)
            
            config = {
                'start_page': self.start_page_var.get().strip(),
                'end_page': self.end_page_var.get().strip(),
                'page_size': self.page_size_var.get().strip()
            }
            
            with open(self.config_file, 'w', encoding='utf-8') as f:
                json.dump(config, f, ensure_ascii=False, indent=2)
            
            self.logger.info("配置已保存")
            messagebox.showinfo("成功", "配置已保存！")
        except Exception as e:
            self.logger.error(f"保存配置失败: {str(e)}")
            messagebox.showerror("错误", f"保存配置失败：{str(e)}")
        
    def setup_logging(self):
        """设置日志处理器"""
        # 创建自定义日志处理器
        class TextHandler(logging.Handler):
            def __init__(self, text_widget):
                logging.Handler.__init__(self)
                self.text_widget = text_widget
                
            def emit(self, record):
                msg = self.format(record)
                def append():
                    self.text_widget.configure(state='normal')
                    self.text_widget.insert(tk.END, msg + '\n')
                    self.text_widget.see(tk.END)
                    self.text_widget.configure(state='disabled')
                self.text_widget.after(0, append)
        
        # 创建独立的logger
        self.logger = logging.getLogger('violation_list')
        self.logger.setLevel(logging.INFO)
        
        # 移除所有现有的处理器
        for handler in self.logger.handlers[:]:
            self.logger.removeHandler(handler)
        
        # 添加文本处理器
        text_handler = TextHandler(self.log_text)
        text_handler.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))
        self.logger.addHandler(text_handler)
        
        # 添加一个初始日志
        self.logger.info("违规商品列表工具已初始化")

    def start_crawling(self):
        """开始爬取数据"""
        # 获取系统配置
        cookie = self.config.get_seller_cookie()
        
        if not cookie:
            messagebox.showerror("错误", "请先在系统配置中设置Cookie")
            return
            
        # 获取分页设置
        try:
            start_page = int(self.start_page_var.get())
            end_page = int(self.end_page_var.get())
            page_size = int(self.page_size_var.get())
            
            if start_page <= 0 or end_page <= 0 or page_size <= 0:
                messagebox.showerror("错误", "页码和每页数量必须大于0")
                return
                
            if start_page > end_page:
                messagebox.showerror("错误", "起始页不能大于结束页")
                return
        except ValueError:
            messagebox.showerror("错误", "请输入有效的数字！")
            return
        
        try:
            # 禁用开始按钮，启用停止按钮
            self.start_button.config(state='disabled')
            self.stop_button.config(state='normal')
            
            # 清空进度条
            self.progress_var.set(0)
            
            # 设置停止标志
            self._stop_flag = False
            
            # 创建爬虫实例
            self.crawler = ViolationListCrawler()
            
            # 在新线程中运行爬虫
            self.crawler_thread = threading.Thread(
                target=self.run_crawler,
                args=(cookie, start_page, end_page, page_size)
            )
            self.crawler_thread.daemon = True
            self.crawler_thread.start()
            
        except Exception as e:
            self.logger.error(f"启动爬虫失败: {str(e)}")
            messagebox.showerror("错误", f"启动爬虫失败: {str(e)}")
            # 恢复按钮状态
            self.start_button.config(state='normal')
            self.stop_button.config(state='disabled')
    
    def run_crawler(self, cookie: str, start_page: int, end_page: int, page_size: int):
        """运行爬虫"""
        try:
            self.logger.info(f"开始获取第 {start_page} 到 {end_page} 页的数据")
            
            # 获取数据
            self.violation_data = self.crawler.get_all_data(
                cookie=cookie,
                start_page=start_page,
                end_page=end_page,
                page_size=page_size,
                progress_callback=self.update_progress,
                stop_flag_callback=lambda: self._stop_flag
            )
            
            if not self.violation_data:
                self.logger.info("未获取到数据！请检查配置信息是否正确。")
                return
            
            # 启用导出按钮
            self.export_button.config(state='normal')
            
            # 在日志中显示成功信息
            self.logger.info(f"成功获取 {len(self.violation_data)} 条数据！")
            
        except Exception as e:
            self.logger.error(f"获取数据时发生错误：{str(e)}")
        finally:
            # 恢复按钮状态
            self.start_button.config(state='normal')
            self.stop_button.config(state='disabled')
    
    def stop_crawling(self):
        """停止爬取"""
        self._stop_flag = True
        self.logger.info("用户请求停止，任务将在当前页面处理完成后停止。")
        messagebox.showinfo("提示", "正在停止，任务将在当前页处理完成后自动结束。")
    
    def update_progress(self, value):
        """更新进度条"""
        self.progress_var.set(value)
    
    def clear_all(self):
        """清空所有内容"""
        self.start_page_var.set("1")
        self.end_page_var.set("1")
        self.page_size_var.set("100")
        
        # 清空日志
        self.log_text.config(state='normal')
        self.log_text.delete(1.0, tk.END)
        self.log_text.config(state='disabled')
        
        # 清空数据
        self.violation_data = []
        
        # 禁用导出按钮
        self.export_button.config(state='disabled')
        
        # 清空进度条
        self.progress_var.set(0)
    
    def export_to_excel(self):
        """导出数据到Excel"""
        if not self.violation_data:
            messagebox.showinfo("提示", "没有数据可导出")
            return
        
        try:
            # 生成文件名
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            default_filename = f"违规商品列表_{timestamp}.xlsx"
            
            # 获取保存目录
            save_dir = filedialog.askdirectory(title="选择保存目录")
            if not save_dir:  # 用户取消选择
                return
            
            # 构建完整的文件路径
            file_path = os.path.join(save_dir, default_filename)
            
            # 导出数据
            if self.excel_exporter.export_to_excel(self.violation_data, file_path):
                self.logger.info(f"数据已成功导出到: {file_path}")
            else:
                self.logger.error("导出数据失败")
                
        except Exception as e:
            self.logger.error(f"导出Excel时出错: {str(e)}")
            messagebox.showerror("错误", f"导出Excel时出错: {str(e)}")

    def merge_excel_files(self):
        """合并违规列表和商品列表Excel文件"""
        try:
            # 选择违规列表Excel文件
            violation_file = filedialog.askopenfilename(
                title="选择违规列表Excel文件",
                filetypes=[("Excel文件", "*.xlsx")]
            )
            if not violation_file:
                return
                
            # 选择商品列表Excel文件
            product_file = filedialog.askopenfilename(
                title="选择商品列表Excel文件",
                filetypes=[("Excel文件", "*.xlsx")]
            )
            if not product_file:
                return
                
            # 选择保存目录
            save_dir = filedialog.askdirectory(title="选择保存目录")
            if not save_dir:
                return
                
            # 生成输出文件名
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_file = os.path.join(save_dir, f"合并数据_{timestamp}.xlsx")
            
            # 读取Excel文件
            from openpyxl import load_workbook
            
            # 读取违规列表
            violation_wb = load_workbook(violation_file)
            violation_ws = violation_wb.active
            
            # 读取商品列表
            product_wb = load_workbook(product_file)
            product_ws = product_wb.active
            
            # 获取表头
            violation_headers = [cell.value for cell in violation_ws[1]]
            product_headers = [cell.value for cell in product_ws[1]]
            
            # 检查必要的列是否存在
            if 'SPUID' not in violation_headers or '商品ID' not in product_headers:
                raise ValueError("Excel文件格式不正确，请确保包含SPUID和商品ID列")
            
            # 获取SPUID和商品ID的列索引
            spuid_col = violation_headers.index('SPUID') + 1
            product_id_col = product_headers.index('商品ID') + 1
            
            # 创建商品ID到行的映射
            product_data = {}
            for row in product_ws.iter_rows(min_row=2):
                product_id = row[product_id_col - 1].value
                if product_id:
                    product_data[product_id] = row
            
            # 创建新的工作簿
            from openpyxl import Workbook
            merged_wb = Workbook()
            merged_ws = merged_wb.active
            
            # 写入合并后的表头
            merged_headers = violation_headers + [h for h in product_headers if h != '商品ID']
            for col, header in enumerate(merged_headers, 1):
                merged_ws.cell(row=1, column=col, value=header)
            
            # 合并数据
            row_num = 2
            for row in violation_ws.iter_rows(min_row=2):
                spuid = row[spuid_col - 1].value
                
                # 写入违规列表数据
                for col, cell in enumerate(row, 1):
                    merged_ws.cell(row=row_num, column=col, value=cell.value)
                
                # 如果找到匹配的商品数据，写入商品列表数据
                if spuid in product_data:
                    product_row = product_data[spuid]
                    col_offset = len(violation_headers)
                    for col, cell in enumerate(product_row, 1):
                        if col != product_id_col:  # 跳过商品ID列
                            merged_ws.cell(row=row_num, column=col_offset + col - 1, value=cell.value)
                
                row_num += 1
            
            # 保存合并后的文件
            merged_wb.save(output_file)
            
            self.logger.info(f"数据已成功合并并保存到: {output_file}")
            messagebox.showinfo("成功", "数据合并完成！")
            
        except Exception as e:
            error_msg = f"合并Excel文件时出错: {str(e)}"
            self.logger.error(error_msg)
            messagebox.showerror("错误", error_msg) 