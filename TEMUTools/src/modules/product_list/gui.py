import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import threading
import json
import os
import logging
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from .crawler import ProductListCrawler
from ..system_config.config import SystemConfig
from ..price_review.config import get_price_threshold

# 模块级常量：商品识别码映射
CODE_MAPPING = {
    "CUSHION": "CU-0608",
    "SLEEVE": "BE-0608",
    "SH": "SH-0608",
    "CB": "CB-0608",
    "Sock": "SO-0608",
    "drawing": "CB-0608",
    "Scarf": "BE-0608",
    "Apron": "AP-0608",
    "bag": "CB-0608", # 双肩包
    "beanie": "BE-0608", # 套头帽
    "workcap": "SH-0608", # 工作帽
    "Handkerchief": "SH-0608", # 方巾
    "WindproofMask": "BE-0608", # 防风骑行面罩
}


def get_product_code(key: str) -> str:
    """根据 `CODE_MAPPING` 智能匹配商品码，忽略大小写和复数形式。

    参数:
        key: SKU编码中用于匹配商品码的关键字

    返回:
        对应的商品识别码，未匹配到则返回“未定义”
    """
    normalized_key = key.lower().rstrip('s')
    for mapping_key, code in CODE_MAPPING.items():
        if normalized_key == mapping_key.lower():
            return code
    return "未定义"

class ProductListTab(ttk.Frame):
    def __init__(self, parent):
        super().__init__(parent)
        self.current_data = []  # 存储当前获取的数据
        self.config = SystemConfig()
        self.setup_ui()
        self.setup_logging()

    def setup_ui(self):
        """设置用户界面"""
        # 创建主框架，添加padding
        main_frame = ttk.Frame(self, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # 第一行：使用Frame来组织控件，确保均匀分布
        first_row_frame = ttk.Frame(main_frame)
        first_row_frame.grid(row=0, column=0, columnspan=5, sticky=(tk.W, tk.E), pady=(0, 10))
        
        # 页数输入
        ttk.Label(first_row_frame, text="获取页数:").pack(side=tk.LEFT)
        self.pages_var = tk.StringVar(value="1")
        self.pages_entry = ttk.Entry(first_row_frame, textvariable=self.pages_var, width=10)
        self.pages_entry.pack(side=tk.LEFT, padx=(0, 20))
        
        # 每页数量输入
        ttk.Label(first_row_frame, text="每页数量:").pack(side=tk.LEFT)
        self.page_size_var = tk.StringVar(value="100")
        self.page_size_entry = ttk.Entry(first_row_frame, textvariable=self.page_size_var, width=10)
        self.page_size_entry.pack(side=tk.LEFT, padx=(0, 20))
        
        # 只爬取在售商品选项
        self.only_on_sale_var = tk.BooleanVar(value=False)
        self.only_on_sale_checkbox = ttk.Checkbutton(
            first_row_frame, 
            text="只爬取在售商品", 
            variable=self.only_on_sale_var,
            command=self.on_checkbox_change
        )
        self.only_on_sale_checkbox.pack(side=tk.LEFT)
        
        # 进度条
        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(main_frame, variable=self.progress_var, maximum=100)
        self.progress_bar.grid(row=1, column=0, columnspan=5, sticky=(tk.W, tk.E), pady=10)
        
        # 日志显示框架
        log_frame = ttk.LabelFrame(main_frame, text="操作日志", padding="5")
        log_frame.grid(row=2, column=0, columnspan=5, sticky=(tk.W, tk.E, tk.N, tk.S), pady=5)
        
        self.log_text = tk.Text(log_frame, height=15, width=100, state='disabled')
        scrollbar = ttk.Scrollbar(log_frame, orient=tk.VERTICAL, command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=scrollbar.set)
        
        self.log_text.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S))
        
        log_frame.columnconfigure(0, weight=1)
        log_frame.rowconfigure(0, weight=1)
        
        # 按钮框架
        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=3, column=0, columnspan=5, pady=10)
        
        self.start_button = ttk.Button(button_frame, text="开始爬取", command=self.start_crawling)
        self.start_button.pack(side=tk.LEFT, padx=5)
        self.stop_button = ttk.Button(button_frame, text="停止", command=self.stop_crawling, state=tk.DISABLED)
        self.stop_button.pack(side=tk.LEFT, padx=5)
        
        # 导出按钮
        export_frame = ttk.LabelFrame(button_frame, text="导出", padding="5")
        export_frame.pack(side=tk.LEFT, padx=5)
        self.export_all_button = ttk.Button(export_frame, text="一键导出所有模板", command=self.export_all_templates, state=tk.DISABLED)
        self.export_all_button.pack(side=tk.LEFT, padx=2)
        ttk.Separator(export_frame, orient='vertical').pack(side=tk.LEFT, padx=5, fill='y')
        self.export_list_button = ttk.Button(export_frame, text="导出商品列表", command=self.export_product_list, state=tk.DISABLED)
        self.export_list_button.pack(side=tk.LEFT, padx=2)
        self.export_code_button = ttk.Button(export_frame, text="导出商品码模板", command=self.export_product_code_template, state=tk.DISABLED)
        self.export_code_button.pack(side=tk.LEFT, padx=2)
        self.export_inventory_button = ttk.Button(export_frame, text="导出库存模板", command=self.export_inventory_template, state=tk.DISABLED)
        self.export_inventory_button.pack(side=tk.LEFT, padx=2)
        
        # 配置网格权重
        self.columnconfigure(0, weight=1)
        self.rowconfigure(0, weight=1)
        main_frame.columnconfigure(0, weight=1)
        main_frame.rowconfigure(2, weight=1)

    def on_checkbox_change(self):
        """复选框状态变化时的处理函数"""
        if self.only_on_sale_var.get():
            self.logger.info("已选择只爬取在售商品")
        else:
            self.logger.info("已选择爬取所有商品")

    def setup_logging(self):
        class TextHandler(logging.Handler):
            def __init__(self, text_widget):
                logging.Handler.__init__(self)
                self.text_widget = text_widget
            def emit(self, record):
                msg = self.format(record)
                def append():
                    self.text_widget.configure(state='normal')
                    self.text_widget.insert(tk.END, msg + '\n')
                    self.text_widget.see(tk.END)
                    self.text_widget.configure(state='disabled')
                self.text_widget.after(0, append)
        self.logger = logging.getLogger('product_list')
        self.logger.setLevel(logging.INFO)
        for handler in self.logger.handlers[:]:
            self.logger.removeHandler(handler)
        text_handler = TextHandler(self.log_text)
        text_handler.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))
        self.logger.addHandler(text_handler)
        self.logger.info("商品列表管理工具已初始化")

    def start_crawling(self):
        try:
            pages = int(self.pages_var.get())
            page_size = int(self.page_size_var.get())
            if pages <= 0 or page_size <= 0:
                messagebox.showerror("错误", "页数和每页数据量必须大于0")
                return
            self.start_button.configure(state='disabled')
            self.stop_button.configure(state='normal')
            self.log_text.delete(1.0, tk.END)
            
            # 获取是否只爬取在售商品的选项
            only_on_sale = self.only_on_sale_var.get()
            
            self.crawler_thread = threading.Thread(
                target=self.run_crawler,
                args=(pages, page_size, only_on_sale)
            )
            self.crawler_thread.start()
        except ValueError:
            messagebox.showerror("错误", "请输入有效的数字")

    def run_crawler(self, pages, page_size, only_on_sale=False):
        try:
            crawler = ProductListCrawler(logger=self.logger)
            
            # 根据选项显示不同的日志信息
            if only_on_sale:
                self.logger.info(f"开始获取在售商品列表数据，计划获取 {pages} 页...")
            else:
                self.logger.info(f"开始获取所有商品列表数据，计划获取 {pages} 页...")
                
            all_data = crawler.get_all_data(max_pages=pages, page_size=page_size, only_on_sale=only_on_sale)
            if all_data:
                self.current_data = all_data
                self.logger.info(f"共获取到 {len(all_data)} 条数据")
                self.export_list_button.configure(state='normal')
                self.export_code_button.configure(state='normal')
                self.export_inventory_button.configure(state='normal')
                self.export_all_button.configure(state='normal')
            else:
                self.current_data = []
                self.logger.warning("未获取到任何数据")
                self.export_list_button.configure(state='disabled')
                self.export_code_button.configure(state='disabled')
                self.export_inventory_button.configure(state='disabled')
                self.export_all_button.configure(state='disabled')
        except Exception as e:
            self.current_data = []
            self.logger.error(f"程序执行出错: {str(e)}")
            self.export_list_button.configure(state='disabled')
            self.export_code_button.configure(state='disabled')
            self.export_inventory_button.configure(state='disabled')
            self.export_all_button.configure(state='disabled')
        finally:
            self.after(0, self.reset_ui)

    def stop_crawling(self):
        self.stop_button.configure(state='disabled')
        logging.info("正在停止爬取...")

    def reset_ui(self):
        self.pages_entry.configure(state='normal')
        self.page_size_entry.configure(state='normal')
        self.start_button.configure(state='normal')
        self.stop_button.configure(state='disabled')
        has_data = bool(self.current_data)
        self.export_all_button.configure(state='normal' if has_data else 'disabled')
        self.export_list_button.configure(state='normal' if has_data else 'disabled')
        self.export_code_button.configure(state='normal' if has_data else 'disabled')
        self.export_inventory_button.configure(state='normal' if has_data else 'disabled')

    def export_product_list(self):
        """导出商品列表到Excel文件"""
        if not self.current_data:
            messagebox.showwarning("警告", "没有可导出的数据")
            return
            
        try:
            # 获取保存路径
            default_filename = f"商品列表_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel文件", "*.xlsx")],
                initialfile=default_filename
            )
            
            if not file_path:  # 用户取消保存
                return
                
            # 准备Excel数据
            excel_data = []
            
            for product in self.current_data:
                # 基本信息
                base_info = {
                    '商品ID': product['productId'],
                    '商品名称': product['productName'],
                    '商品类型': product['productType'],
                    '来源类型': product['sourceType'],
                    '商品编码': product['goodsId'],
                    '主图URL': product['mainImageUrl'],
                    '创建时间': datetime.fromtimestamp(product['createdAt']/1000).strftime('%Y-%m-%d %H:%M:%S'),
                    '类目ID': product['leafCat']['catId'],
                    '类目名称': product['leafCat']['catName']
                }
                
                # 处理SKU信息
                for sku in product['productSkuSummaries']:
                    sku_info = base_info.copy()
                    supplier_price = sku['supplierPrice'] / 100  # 转换为元
                    
                    # 计算底线价格和价格差额
                    sku_code = sku['extCode']
                    price_threshold = get_price_threshold(sku_code)
                    price_difference = supplier_price - price_threshold if price_threshold is not None else None
                    
                    sku_info.update({
                        'SKU ID': sku['productSkuId'],
                        'SKU编码': sku['extCode'],
                        '供应商价格': supplier_price,
                        '底线价格': price_threshold,
                        '价格差额': price_difference,
                        'SKU图片': sku['thumbUrl']
                    })
                    
                    # 处理规格信息
                    for spec in sku['productSkuSpecList']:
                        sku_info[f'{spec["parentSpecName"]}'] = spec['specName']
                    
                    excel_data.append(sku_info)
            
            # 创建工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "商品列表"
            
            # 设置样式
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 设置列顺序
            columns_order = [
                '商品ID', '商品名称', '商品类型', '来源类型', '商品编码',
                '类目ID', '类目名称', 'SKU ID', 'SKU编码', '供应商价格',
                '底线价格', '价格差额', '颜色', '尺码', '主图URL', 'SKU图片', '创建时间'
            ]
            
            # 写入表头
            for col, header in enumerate(columns_order, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
                # 设置列宽
                ws.column_dimensions[get_column_letter(col)].width = 20
            
            # 写入数据
            for row, data in enumerate(excel_data, 2):
                for col, header in enumerate(columns_order, 1):
                    value = data.get(header, '')
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                    cell.border = border
            
            # 保存文件
            wb.save(file_path)
            
            logging.info(f"商品列表已导出到Excel: {file_path}")
            messagebox.showinfo("成功", "商品列表导出成功！")
            
        except Exception as e:
            error_msg = f"导出商品列表失败: {str(e)}"
            logging.error(error_msg)
            messagebox.showerror("错误", error_msg)

    def export_product_code_template(self):
        """导出商品码模板"""
        if not self.current_data:
            messagebox.showwarning("警告", "没有可导出的数据")
            return
            
        try:
            # 获取保存路径
            default_filename = f"商品码模板_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel文件", "*.xlsx")],
                initialfile=default_filename
            )
            
            if not file_path:  # 用户取消保存
                return
                
            # 准备Excel数据
            excel_data = []
            
            for product in self.current_data:
                product_id = product['productId']
                
                # 处理SKU信息
                for sku in product['productSkuSummaries']:
                    sku_code = sku['extCode']
                    # 获取下划线前的部分
                    key = sku_code.split('_')[0] if '_' in sku_code else sku_code
                    
                    # 使用模块级智能匹配函数查找对应的商品码
                    product_code = get_product_code(key)
                    
                    excel_data.append({
                        'SPUID': product_id,
                        '商品识别码': product_code
                    })
            
            # 创建工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "商品码模板"
            
            # 设置样式
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 写入表头
            headers = ['SPUID', '商品识别码']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=f"{header}")
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
                # 设置列宽
                ws.column_dimensions[get_column_letter(col)].width = 20
            
            # 写入数据
            for row, data in enumerate(excel_data, 2):
                for col, header in enumerate(headers, 1):
                    value = data.get(header, '')
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    cell.border = border
            
            # 保存文件
            wb.save(file_path)
            
            logging.info(f"商品码模板已导出到Excel: {file_path}")
            messagebox.showinfo("成功", "商品码模板导出成功！")
            
        except Exception as e:
            error_msg = f"导出商品码模板失败: {str(e)}"
            logging.error(error_msg)
            messagebox.showerror("错误", error_msg)

    def export_inventory_template(self):
        """导出库存模板"""
        if not self.current_data:
            messagebox.showwarning("警告", "没有可导出的数据")
            return
            
        try:
            # 收集所有SKU ID
            all_sku_ids = []
            for product in self.current_data:
                for sku in product['productSkuSummaries']:
                    all_sku_ids.append(sku['productSkuId'])
            
            if not all_sku_ids:
                messagebox.showwarning("警告", "没有找到SKU数据")
                return
            
            # 计算需要拆分的文件数量
            max_records_per_file = 1000
            total_sku_count = len(all_sku_ids)
            file_count = (total_sku_count + max_records_per_file - 1) // max_records_per_file
            
            if file_count == 1:
                # 单个文件，使用原来的逻辑
                default_filename = f"库存模板_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                file_path = filedialog.asksaveasfilename(
                    defaultextension=".xlsx",
                    filetypes=[("Excel文件", "*.xlsx")],
                    initialfile=default_filename
                )
                
                if not file_path:  # 用户取消保存
                    return
                    
                self.export_inventory_template_to_path(file_path, all_sku_ids)
                logging.info(f"库存模板已导出到Excel: {file_path}")
                messagebox.showinfo("成功", "库存模板导出成功！")
                
            else:
                # 多个文件，选择保存目录
                save_dir = filedialog.askdirectory(title="选择保存目录")
                if not save_dir:  # 用户取消选择
                    return
                
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                
                # 拆分并导出多个文件
                for i in range(file_count):
                    start_idx = i * max_records_per_file
                    end_idx = min((i + 1) * max_records_per_file, total_sku_count)
                    current_sku_ids = all_sku_ids[start_idx:end_idx]
                    
                    # 生成文件名
                    if file_count == 1:
                        filename = f"库存模板_{timestamp}.xlsx"
                    else:
                        filename = f"库存模板_{timestamp}-{i+1}.xlsx"
                    
                    file_path = os.path.join(save_dir, filename)
                    self.export_inventory_template_to_path(file_path, current_sku_ids)
                
                logging.info(f"库存模板已拆分为 {file_count} 个文件并导出到目录: {save_dir}")
                messagebox.showinfo("成功", f"库存模板已拆分为 {file_count} 个文件导出成功！\n每个文件最多包含 {max_records_per_file} 条记录。")
            
        except Exception as e:
            error_msg = f"导出库存模板失败: {str(e)}"
            logging.error(error_msg)
            messagebox.showerror("错误", error_msg)

    def export_inventory_template_to_path(self, file_path, sku_ids=None):
        """导出库存模板到指定路径"""
        try:
            # 如果没有提供SKU ID列表，则从当前数据中收集
            if sku_ids is None:
                sku_ids = []
                for product in self.current_data:
                    for sku in product['productSkuSummaries']:
                        sku_ids.append(sku['productSkuId'])
            
            # 创建工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "库存模板"
            
            # 设置样式
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 写入表头
            headers = ['SKU ID', '', '', '修改后库存']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
                # 设置列宽
                ws.column_dimensions[get_column_letter(col)].width = 20
            
            # 写入说明行
            ws.cell(row=2, column=1, value="必填指Temu的SKU ID")
            ws.cell(row=2, column=4, value="条件必填指当前的实际总库存，导入成功后将直接覆盖线上已有库存")
            
            # 写入数据
            for row, sku_id in enumerate(sku_ids, 3):
                # SKU ID (第1列)
                cell = ws.cell(row=row, column=1, value=sku_id)
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.border = border
                
                # 第2、3列保持为空
                for col in [2, 3]:
                    cell = ws.cell(row=row, column=col, value="")
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    cell.border = border
                
                # 默认库存值 (第4列)
                cell = ws.cell(row=row, column=4, value=1000)
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.border = border
            
            # 保存文件
            wb.save(file_path)
            
        except Exception as e:
            raise Exception(f"导出库存模板失败: {str(e)}")

    def export_all_templates(self):
        """一键导出所有模板"""
        if not self.current_data:
            messagebox.showwarning("警告", "没有可导出的数据")
            return
            
        try:
            # 获取保存目录
            save_dir = filedialog.askdirectory(title="选择保存目录")
            if not save_dir:  # 用户取消选择
                return
                
            # 生成时间戳
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            
            # 导出商品列表
            list_path = os.path.join(save_dir, f"商品列表_{timestamp}.xlsx")
            self.export_product_list_to_path(list_path)
            
            # 导出商品码模板
            code_path = os.path.join(save_dir, f"商品码模板_{timestamp}.xlsx")
            self.export_product_code_template_to_path(code_path)
            
            # 导出库存模板（支持拆分）
            all_sku_ids = []
            for product in self.current_data:
                for sku in product['productSkuSummaries']:
                    all_sku_ids.append(sku['productSkuId'])
            
            if all_sku_ids:
                # 计算需要拆分的文件数量
                max_records_per_file = 1000
                total_sku_count = len(all_sku_ids)
                file_count = (total_sku_count + max_records_per_file - 1) // max_records_per_file
                
                # 拆分并导出多个库存模板文件
                for i in range(file_count):
                    start_idx = i * max_records_per_file
                    end_idx = min((i + 1) * max_records_per_file, total_sku_count)
                    current_sku_ids = all_sku_ids[start_idx:end_idx]
                    
                    # 生成文件名
                    if file_count == 1:
                        filename = f"库存模板_{timestamp}.xlsx"
                    else:
                        filename = f"库存模板_{timestamp}-{i+1}.xlsx"
                    
                    inventory_path = os.path.join(save_dir, filename)
                    self.export_inventory_template_to_path(inventory_path, current_sku_ids)
                
                logging.info(f"所有模板已导出到目录: {save_dir}")
                if file_count > 1:
                    messagebox.showinfo("成功", f"所有模板导出成功！\n库存模板已拆分为 {file_count} 个文件，每个文件最多包含 {max_records_per_file} 条记录。")
                else:
                    messagebox.showinfo("成功", "所有模板导出成功！")
            else:
                logging.info(f"所有模板已导出到目录: {save_dir}")
                messagebox.showinfo("成功", "所有模板导出成功！")
            
        except Exception as e:
            error_msg = f"导出模板失败: {str(e)}"
            logging.error(error_msg)
            messagebox.showerror("错误", error_msg)

    def export_product_list_to_path(self, file_path):
        """导出商品列表到指定路径"""
        try:
            # 准备Excel数据
            excel_data = []
            
            for product in self.current_data:
                # 基本信息
                base_info = {
                    '商品ID': product['productId'],
                    '商品名称': product['productName'],
                    '商品类型': product['productType'],
                    '来源类型': product['sourceType'],
                    '商品编码': product['goodsId'],
                    '主图URL': product['mainImageUrl'],
                    '创建时间': datetime.fromtimestamp(product['createdAt']/1000).strftime('%Y-%m-%d %H:%M:%S'),
                    '类目ID': product['leafCat']['catId'],
                    '类目名称': product['leafCat']['catName']
                }
                
                # 处理SKU信息
                for sku in product['productSkuSummaries']:
                    sku_info = base_info.copy()
                    supplier_price = sku['supplierPrice'] / 100  # 转换为元
                    
                    # 计算底线价格和价格差额
                    sku_code = sku['extCode']
                    price_threshold = get_price_threshold(sku_code)
                    price_difference = supplier_price - price_threshold if price_threshold is not None else None
                    
                    sku_info.update({
                        'SKU ID': sku['productSkuId'],
                        'SKU编码': sku['extCode'],
                        '供应商价格': supplier_price,
                        '底线价格': price_threshold,
                        '价格差额': price_difference,
                        'SKU图片': sku['thumbUrl']
                    })
                    
                    # 处理规格信息
                    for spec in sku['productSkuSpecList']:
                        sku_info[f'{spec["parentSpecName"]}'] = spec['specName']
                    
                    excel_data.append(sku_info)
            
            # 创建工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "商品列表"
            
            # 设置样式
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 设置列顺序
            columns_order = [
                '商品ID', '商品名称', '商品类型', '来源类型', '商品编码',
                '类目ID', '类目名称', 'SKU ID', 'SKU编码', '供应商价格',
                '底线价格', '价格差额', '颜色', '尺码', '主图URL', 'SKU图片', '创建时间'
            ]
            
            # 写入表头
            for col, header in enumerate(columns_order, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
                # 设置列宽
                ws.column_dimensions[get_column_letter(col)].width = 20
            
            # 写入数据
            for row, data in enumerate(excel_data, 2):
                for col, header in enumerate(columns_order, 1):
                    value = data.get(header, '')
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                    cell.border = border
            
            # 保存文件
            wb.save(file_path)
            
        except Exception as e:
            raise Exception(f"导出商品列表失败: {str(e)}")

    def export_product_code_template_to_path(self, file_path):
        """导出商品码模板到指定路径"""
        try:
            # 准备Excel数据
            excel_data = []
            
            for product in self.current_data:
                product_id = product['productId']
                
                # 处理SKU信息
                for sku in product['productSkuSummaries']:
                    sku_code = sku['extCode']
                    # 获取下划线前的部分
                    key = sku_code.split('_')[0] if '_' in sku_code else sku_code
                    
                    # 使用模块级智能匹配函数查找对应的商品码
                    product_code = get_product_code(key)
                    
                    excel_data.append({
                        'SPUID': product_id,
                        '商品识别码': product_code
                    })
            
            # 创建工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "商品码模板"
            
            # 设置样式
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 写入表头
            headers = ['SPUID', '商品识别码']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=f"{header}")
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
                # 设置列宽
                ws.column_dimensions[get_column_letter(col)].width = 20
            
            # 写入数据
            for row, data in enumerate(excel_data, 2):
                for col, header in enumerate(headers, 1):
                    value = data.get(header, '')
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    cell.border = border
            
            # 保存文件
            wb.save(file_path)
            
        except Exception as e:
            raise Exception(f"导出商品码模板失败: {str(e)}") 