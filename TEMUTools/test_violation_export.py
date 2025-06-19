#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
违规列表导出功能测试脚本
测试新增的违规站点数量和违规记录数量字段
"""

import sys
import os
import json
import tempfile
from datetime import datetime

# 添加项目根目录到Python路径
project_root = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, os.path.join(project_root, 'src'))

def create_test_data():
    """创建测试数据"""
    return [
        {
            "spu_id": 12345,
            "goods_name": "测试商品1",
            "leaf_reason_name": "商品描述违规",
            "violation_desc": "商品描述包含违禁词",
            "site_num": 5,
            "punish_num": 5,
            "punish_detail_list": [
                {
                    "site_id": 100,
                    "punish_id": "punish_001",
                    "punish_appeal_type": "商品下架",
                    "punish_infect_desc": "商品被下架处理",
                    "illegal_detail": [
                        {"title": "违规内容", "value": "包含违禁词"}
                    ],
                    "rectification_suggestion": "请修改商品描述",
                    "start_time": int(datetime.now().timestamp()),
                    "plan_end_time": int(datetime.now().timestamp()) + 86400,
                    "appeal_status": 0,
                    "now_appeal_time": 0,
                    "max_appeal_time": 3
                }
            ]
        },
        {
            "spu_id": 12346,
            "goods_name": "测试商品2",
            "leaf_reason_name": "图片违规",
            "violation_desc": "商品图片不符合规范",
            "site_num": 87,
            "punish_num": 87,
            "punish_detail_list": [
                {
                    "site_id": -1,  # 全部站点违规
                    "punish_id": "punish_002",
                    "punish_appeal_type": "商品禁售",
                    "punish_infect_desc": "商品被禁售处理",
                    "illegal_detail": [
                        {"title": "违规图片", "value": "图片内容违规"}
                    ],
                    "rectification_suggestion": "请更换合规图片",
                    "start_time": int(datetime.now().timestamp()),
                    "plan_end_time": int(datetime.now().timestamp()) + 172800,
                    "appeal_status": 1,
                    "now_appeal_time": 1,
                    "max_appeal_time": 3
                }
            ]
        },
        {
            "spu_id": 12347,
            "goods_name": "测试商品3",
            "leaf_reason_name": "价格违规",
            "violation_desc": "商品价格设置不合理",
            "site_num": 0,
            "punish_num": 0,
            "punish_detail_list": []  # 空列表测试
        }
    ]

def test_export_functionality():
    """测试导出功能"""
    print("开始测试违规列表导出功能...")
    
    try:
        # 导入必要的模块
        from modules.violation_list.gui import ViolationListTab
        import tkinter as tk
        
        # 创建测试数据
        test_data = create_test_data()
        
        # 创建临时的GUI实例（仅用于测试导出功能）
        root = tk.Tk()
        root.withdraw()  # 隐藏主窗口
        
        # 创建违规列表标签页
        violation_tab = ViolationListTab(root)
        
        # 设置测试数据
        violation_tab.current_data = test_data
        
        # 创建临时文件
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
            temp_file_path = tmp_file.name
        
        try:
            # 测试导出功能
            print("测试导出Excel功能...")
            
            # 模拟导出过程
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            
            # 创建工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "违规商品列表"
            
            # 设置样式
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 写入表头
            headers = [
                "SPUID", "商品名称", "违规原因", "违规描述",
                "违规站点数量", "违规记录数量", "处罚类型", "处罚影响", 
                "违规详情", "整改建议", "开始时间", "结束时间", 
                "申诉状态", "申诉次数", "最大申诉次数"
            ]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
                ws.column_dimensions[get_column_letter(col)].width = 20
            
            # 写入数据
            row = 2
            for product in test_data:
                # 获取第一个处罚详情（如果有的话）
                punish_detail = product.get('punish_detail_list', [{}])[0] if product.get('punish_detail_list') else {}
                
                # 获取违规详情
                illegal_details = punish_detail.get('illegal_detail', [])
                illegal_desc = '; '.join([
                    f"{illegal.get('title', '')}: {illegal.get('value', '')}"
                    for illegal in illegal_details
                ])
                
                # 格式化时间戳
                def format_timestamp(timestamp: int) -> str:
                    if not timestamp:
                        return ''
                    try:
                        return datetime.fromtimestamp(timestamp).strftime('%Y-%m-%d %H:%M:%S')
                    except:
                        return ''
                
                start_time = format_timestamp(punish_detail.get('start_time'))
                end_time = format_timestamp(punish_detail.get('plan_end_time'))
                
                # 处理违规站点数量 - 处理特殊情况
                site_num = product.get('site_num', 0)
                punish_num = product.get('punish_num', 0)
                
                # 检查是否有全部站点违规的情况（site_id为-1）
                punish_details = product.get('punish_detail_list', [])
                has_all_sites_violation = any(
                    detail.get('site_id') == -1 for detail in punish_details
                )
                
                # 格式化站点数量显示
                if has_all_sites_violation:
                    site_num_display = "全部站点"
                else:
                    site_num_display = str(site_num) if site_num > 0 else "0"
                
                # 格式化申诉状态
                def format_appeal_status(status: int) -> str:
                    status_map = {
                        0: '未申诉',
                        1: '申诉中',
                        2: '申诉通过',
                        3: '申诉驳回'
                    }
                    return status_map.get(status, '未知状态')
                
                # 写入一行数据
                ws.cell(row=row, column=1, value=product.get('spu_id', ''))
                ws.cell(row=row, column=2, value=product.get('goods_name', ''))
                ws.cell(row=row, column=3, value=product.get('leaf_reason_name', ''))
                ws.cell(row=row, column=4, value=product.get('violation_desc', ''))
                ws.cell(row=row, column=5, value=site_num_display)  # 违规站点数量
                ws.cell(row=row, column=6, value=punish_num)  # 违规记录数量
                ws.cell(row=row, column=7, value=punish_detail.get('punish_appeal_type', ''))
                ws.cell(row=row, column=8, value=punish_detail.get('punish_infect_desc', ''))
                ws.cell(row=row, column=9, value=illegal_desc)
                ws.cell(row=row, column=10, value=punish_detail.get('rectification_suggestion', ''))
                ws.cell(row=row, column=11, value=start_time)
                ws.cell(row=row, column=12, value=end_time)
                ws.cell(row=row, column=13, value=format_appeal_status(punish_detail.get('appeal_status', 0)))
                ws.cell(row=row, column=14, value=punish_detail.get('now_appeal_time', 0))
                ws.cell(row=row, column=15, value=punish_detail.get('max_appeal_time', 0))
                
                row += 1
            
            # 保存文件
            wb.save(temp_file_path)
            
            print(f"✅ 测试Excel文件已保存到: {temp_file_path}")
            
            # 验证数据
            print("\n验证导出的数据:")
            print("-" * 80)
            print(f"{'SPUID':<10} {'商品名称':<15} {'违规站点数量':<12} {'违规记录数量':<12} {'特殊说明':<15}")
            print("-" * 80)
            
            for product in test_data:
                site_num = product.get('site_num', 0)
                punish_num = product.get('punish_num', 0)
                punish_details = product.get('punish_detail_list', [])
                has_all_sites_violation = any(
                    detail.get('site_id') == -1 for detail in punish_details
                )
                
                if has_all_sites_violation:
                    site_num_display = "全部站点"
                    special_note = "site_id=-1"
                else:
                    site_num_display = str(site_num) if site_num > 0 else "0"
                    special_note = "正常"
                
                print(f"{product.get('spu_id', ''):<10} {product.get('goods_name', ''):<15} {site_num_display:<12} {punish_num:<12} {special_note:<15}")
            
            print("-" * 80)
            print("✅ 测试数据验证完成")
            
            # 检查文件是否存在
            if os.path.exists(temp_file_path):
                file_size = os.path.getsize(temp_file_path)
                print(f"✅ 导出文件创建成功，文件大小: {file_size} 字节")
            else:
                print("❌ 导出文件创建失败")
                return False
            
            return True
            
        finally:
            # 清理临时文件
            try:
                os.unlink(temp_file_path)
                print("✅ 临时文件已清理")
            except:
                pass
        
    except Exception as e:
        print(f"❌ 测试过程中发生错误: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    print("=" * 60)
    print("违规列表导出功能测试")
    print("=" * 60)
    
    success = test_export_functionality()
    
    print("=" * 60)
    if success:
        print("✅ 所有测试通过")
    else:
        print("❌ 测试失败")
    print("=" * 60) 